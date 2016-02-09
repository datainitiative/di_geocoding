# django imports
from django.shortcuts import render_to_response, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.template import RequestContext
from django.db import models
from django.db.models.loading import get_model
from django.contrib import messages
from django.contrib.auth import authenticate,login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm, SetPasswordForm
from django.core.files import File
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

# python imports
from zipfile import *
import collections # for OrderedDict

# Import from general utilities
from util import *

# Import from app
from geocodingtool.settings import ROOT_APP_URL, STORAGE_ROOTPATH, STATIC_URL
from geocodingtool.settings import GOOGLE_API_KEY, OSM_API_KEY, ARCGIS_API_KEY, BING_API_KEY, MAPBOX_API_KEY, MAPQUEST_API_KEY, GOOGLE_API_LIMIT
from geocodingtool.settings import EXCEL_FILE_EXTENSIONS, CSV_FILE_EXTENSIONS
from geocodingtool.settings import CF_GEOTABLE
from geocodingapp.models import *
from geocodingapp.forms import *

# Import from 3rd party libs
## Geocoder
import geocoder
## Import xlrd for reading XLS files with Python
from openpyxl import load_workbook
from openpyxl.utils import (_get_column_letter)

'''-----------------------
Home Page
-----------------------'''
def update_all_geocoders_usage():
    geocoders = Geocoder.objects.all()
    geocoder_status = []
    for geocoder in geocoders:
        if geocoder.limit == -1:
            geocoder_status.append({'name':geocoder.name,'limit':"Unlimited",'limit_unit':""})
        elif geocoder.limit > -1:
            if geocoder.name == "Google Maps":
                try:
                    geocoder_usages = GeocoderUsage.objects.filter(geocoder=geocoder).filter(has_expired=False)
                    total_limit = GOOGLE_API_LIMIT
                    timenow = datetime.datetime.utcnow()
                    for gu in geocoder_usages:
                        last_geocoding_time = gu.last_geocoding_time.replace(tzinfo=None) # remove time zone info                     
                        timedelta_in_day = (timenow-last_geocoding_time).days
                        if timedelta_in_day >= 1:
                            gu.has_expired = True
                            gu.save()
                        else:
                            total_limit -= gu.geocoding_record_num
                    geocoder_status.append({'name':geocoder.name,'limit':str(total_limit),'limit_unit':geocoder.limit_unit})
                except Exception as e:
                    print e
            elif geocoder.name == "Bing Maps":
                timenow = datetime.datetime.now()
                today_date = "%s/%s" % (timenow.month,timenow.day)
                if today_date == "1/1":
                    geocoder.limit = 125000
                    geocoder.save()
                geocoder_status.append({'name':geocoder.name,'limit':str(geocoder.limit),'limit_unit':geocoder.limit_unit})
            elif geocoder.name == "Mapquest":
                timenow = datetime.datetime.now()
                today_date = "%s" % timenow.day
                if today_date == "1":
                    geocoder.limit = 15000
                    geocoder.save()
                geocoder_status.append({'name':geocoder.name,'limit':str(geocoder.limit),'limit_unit':geocoder.limit_unit})
            else:
                geocoder_status.append({'name':geocoder.name,'limit':str(geocoder.limit),'limit_unit':geocoder.limit_unit})
    return geocoder_status

# Home page
@login_required
@render_to("geocodingapp/home.html")
def home(request):
    projects = Project.objects.all()
    num_project = len(projects)
    tasks = Task.objects.all()
    num_task = len(tasks)
    geocoder_status = update_all_geocoders_usage()
       
    return {
        "num_project": num_project,
        "num_task": num_task,
        "geocoder_status": geocoder_status,
    }


'''-----------------------
Functions
-----------------------'''
def update_google_geocoder_usage(usage_id=None,record_num=0,new_usage=True):
    if new_usage:
        google_geocoder_usage = GeocoderUsage(
                geocoder = Geocoder.objects.get(name="Google Maps"),
                geocoding_record_num = record_num
            )
    else:
        google_geocoder_usage = GeocoderUsage.objects.get(id=usage_id)
        google_geocoder_usage.geocoding_record_num += 1
    google_geocoder_usage.save()
    return google_geocoder_usage

def substract_geocoder_usage(geocoder_name=None):
    if geocoder_name and (geocoder_name != ""):
        geocoder = Geocoder.objects.get(name=geocoder_name)
        geocoder.limit -= 1
        geocoder.save()

def api_geocoding(address,usage_id=None):
    t1 = time.time()
    
    confidence_level = -1
    highest_confidence_level = -1
    min_bbox = None
    best_geocoder = None
    best_g_latlng = None
    best_accuracy = None
    best_formatted_address = None
    is_fail = False
    # Google geocoder
    print address
    print "Start geocoding, try Google"
    # update google usage
    if usage_id:
        update_google_geocoder_usage(usage_id,1,False)    
    g = geocoder.google(address)
    print g.geojson
    if g.geojson['properties']['ok']:
        try:
            confidence_level = g.geojson['properties']['confidence']
        except:
            confidence_level = 0
        # There is a change in geojson return, add a condition here
        # >>> g = geocoder.google("address")
        # >>> g.bbox
        # {"northeast": [lat1, lng1], "southwest": [lat2, lng2]}
        # >>> g.geojson['bbox']
        # [lng2, lat2, lng1, lat1]
        # >>> g.southwest
        # [lat2, lng1]
        if ('northeast' in g.geojson['bbox']) and ('southwest' in g.geojson['bbox']):
            bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
        else:
            bbox = round(math.sqrt((g.geojson['bbox'][3] - g.geojson['bbox'][1])**2 + (g.geojson['bbox'][2] - g.geojson['bbox'][0])**2),8)
        if confidence_level > highest_confidence_level:
            highest_confidence_level = confidence_level
            min_bbox = bbox
            print "bbox: ", bbox
            g_latlng = g.geojson['geometry']['coordinates']
            g_latlng.reverse()
            best_g_latlng = g_latlng
            best_geocoder = "Google Maps"
            try:
                best_accuracy = g.geojson['properties']['accuracy']
            except:
                best_accuracy = ""
            try:
                best_formatted_address = g.geojson['properties']['address']
            except:
                best_formatted_address = ""
            if confidence_level < 8:
                is_fail = True
            else:
                is_fail = False
        else:
            is_fail = True
    else:
        is_fail = True
    if is_fail:        
        print "confidence level: ", confidence_level
        print "google failed, try osm"
        print "---------------"
        confidence_level = None
        bbox = None            
        # OSM geocoder
        g = geocoder.osm(address)
        print g.geojson
        if g.geojson['properties']['ok']:
            try:
                confidence_level = g.geojson['properties']['confidence']
            except:
                confidence_level = 0                
            confidence_level = g.geojson['properties']['confidence']
            if ('northeast' in g.geojson['bbox']) and ('southwest' in g.geojson['bbox']):
                bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
            else:
                bbox = round(math.sqrt((g.geojson['bbox'][3] - g.geojson['bbox'][1])**2 + (g.geojson['bbox'][2] - g.geojson['bbox'][0])**2),8)            
            if confidence_level >= highest_confidence_level:
                if confidence_level == highest_confidence_level:
                    print "bbox: ", bbox
                    if (not min_bbox) or (bbox < min_bbox):
                        min_bbox = bbox
                        highest_confidence_level = confidence_level
                        g_latlng = g.geojson['geometry']['coordinates']
                        g_latlng.reverse()
                        best_g_latlng = g_latlng
                        best_geocoder = "OpenStreetMaps"
                    try:
                        best_accuracy = g.geojson['properties']['accuracy']
                    except:
                        best_accuracy = ""
                    try:
                        best_formatted_address = g.geojson['properties']['address']
                    except:
                        best_formatted_address = ""                    
                    else:
                        is_fail = True
                else:
                    min_bbox = bbox
                    highest_confidence_level = confidence_level
                    g_latlng = g.geojson['geometry']['coordinates']
                    g_latlng.reverse()
                    best_g_latlng = g_latlng
                    best_geocoder = "OpenStreetMaps"
                    try:
                        best_accuracy = g.geojson['properties']['accuracy']
                    except:
                        best_accuracy = ""    
                    try:
                        best_formatted_address = g.geojson['properties']['address']
                    except:
                        best_formatted_address = ""                    
                if confidence_level < 8:
                    is_fail = True
                else:
                    is_fail = False
            else:
                is_fail = True                
        else:              
            is_fail = True
        if is_fail:               
            print "confidence level: ", confidence_level
            print "osm failed, try arcgis"
            print "---------------"   
            confidence_level = None
            bbox = None                              
            # ArcGIS geocoder
            g = geocoder.arcgis(address)
            print g.geojson                
            if g.geojson['properties']['ok']:
                try:
                    confidence_level = g.geojson['properties']['confidence']
                except:
                    confidence_level = 0                    
                confidence_level = g.geojson['properties']['confidence']
                if ('northeast' in g.geojson['bbox']) and ('southwest' in g.geojson['bbox']):
                    bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                else:
                    bbox = round(math.sqrt((g.geojson['bbox'][3] - g.geojson['bbox'][1])**2 + (g.geojson['bbox'][2] - g.geojson['bbox'][0])**2),8)                
                if confidence_level >= highest_confidence_level:
                    if confidence_level == highest_confidence_level:
                        print "bbox: ", bbox
                        if (min_bbox) or (bbox > min_bbox):
                            min_bbox = bbox
                            highest_confidence_level = confidence_level
                            g_latlng = g.geojson['geometry']['coordinates']
                            g_latlng.reverse()
                            best_g_latlng = g_latlng
                            best_geocoder = "ArcGIS"
                            try:
                                best_accuracy = g.geojson['properties']['accuracy']
                            except:
                                best_accuracy = ""                            
                            try:
                                best_formatted_address = g.geojson['properties']['address']
                            except:
                                best_formatted_address = ""                            
                        else:
                            is_fail = True
                    else:                        
                        highest_confidence_level = confidence_level
                        g_latlng = g.geojson['geometry']['coordinates']
                        g_latlng.reverse()
                        best_g_latlng = g_latlng
                        best_geocoder = "ArcGIS"
                        try:
                            best_accuracy = g.geojson['properties']['accuracy']
                        except:
                            best_accuracy = ""
                        try:
                            best_formatted_address = g.geojson['properties']['address']
                        except:
                            best_formatted_address = ""                        
                    if confidence_level < 8:
                        is_fail = True
                    else:
                        is_fail = False
                else:
                    is_fail = True                    
            else:                  
                is_fail = True
            if is_fail:                   
                print "confidence level: ", confidence_level
                print "arcgis failed, try bing"
                print "---------------"
                confidence_level = None
                bbox = None                     
                # Bing geocoder
                g = geocoder.bing(address,key=BING_API_KEY)
                substract_geocoder_usage("Bing Maps")
                print g.geojson
                if g.geojson['properties']['ok']:
                    try:
                        confidence_level = g.geojson['properties']['confidence']
                    except:
                        confidence_level = 0                        
                    confidence_level = g.geojson['properties']['confidence']
                    if ('northeast' in g.geojson['bbox']) and ('southwest' in g.geojson['bbox']):
                        bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                    else:
                        bbox = round(math.sqrt((g.geojson['bbox'][3] - g.geojson['bbox'][1])**2 + (g.geojson['bbox'][2] - g.geojson['bbox'][0])**2),8)                    
                    if confidence_level >= highest_confidence_level:
                        if confidence_level == highest_confidence_level:
                            print "bbox: ", bbox
                            if (min_bbox) or (bbox > min_bbox):
                                min_bbox = bbox
                                highest_confidence_level = confidence_level
                                g_latlng = g.geojson['geometry']['coordinates']
                                g_latlng.reverse()
                                best_g_latlng = g_latlng
                                best_geocoder = "Bing Maps"
                                try:
                                    best_accuracy = g.geojson['properties']['accuracy']
                                except:
                                    best_accuracy = ""
                                try:
                                    best_formatted_address = g.geojson['properties']['address']
                                except:
                                    best_formatted_address = ""                                
                            else:
                                is_fail = True
                        else:                            
                            highest_confidence_level = confidence_level
                            g_latlng = g.geojson['geometry']['coordinates']
                            g_latlng.reverse()
                            best_g_latlng = g_latlng
                            best_geocoder = "Bing Maps"
                            bing_geocoder = Geocoder.objects.get(name="Bing Maps")
                            bing_geocoder.limit -= 1
                            try:
                                best_accuracy = g.geojson['properties']['accuracy']
                            except:
                                best_accuracy = ""
                            try:
                                best_formatted_address = g.geojson['properties']['address']
                            except:
                                best_formatted_address = ""                            
                        if confidence_level < 8:
                            is_fail = True
                        else:
                            is_fail = False
                    else:
                        is_fail = True                        
                else:                      
                    is_fail = True
                if is_fail:                       
                    print "confidence level: ", confidence_level
                    print "bing failed, try mapbox"
                    print "---------------"
                    confidence_level = None
                    bbox = None                         
                    # Mapbox geocoder
                    g = geocoder.mapbox(address,key=MAPBOX_API_KEY)
                    print "bbox: ", bbox
                    print g.geojson
                    if g.geojson['properties']['ok']:
                        try:
                            confidence_level = g.geojson['properties']['confidence']
                        except:
                            confidence_level = 0
                        try:
                            if ('northeast' in g.geojson['bbox']) and ('southwest' in g.geojson['bbox']):
                                bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                            else:
                                bbox = round(math.sqrt((g.geojson['bbox'][3] - g.geojson['bbox'][1])**2 + (g.geojson['bbox'][2] - g.geojson['bbox'][0])**2),8)                            
                        except:
                            bbox = None
                        if confidence_level >= highest_confidence_level:
                            if confidence_level == highest_confidence_level:
                                print "bbox: ", bbox
                                if (min_bbox) or (bbox > min_bbox):
                                    min_bbox = bbox
                                    highest_confidence_level = confidence_level
                                    g_latlng = g.geojson['geometry']['coordinates']
                                    g_latlng.reverse()
                                    best_g_latlng = g_latlng
                                    best_geocoder = "Mapbox"
                                    try:
                                        best_accuracy = g.geojson['properties']['accuracy']
                                    except:
                                        best_accuracy = ""
                                    try:
                                        best_formatted_address = g.geojson['properties']['address']
                                    except:
                                        best_formatted_address = ""                                    
                                else:
                                    is_fail = True
                            else:                                
                                highest_confidence_level = confidence_level
                                g_latlng = g.geojson['geometry']['coordinates']
                                g_latlng.reverse()
                                best_g_latlng = g_latlng
                                best_geocoder = "Mapbox"
                                try:
                                    best_accuracy = g.geojson['properties']['accuracy']
                                except:
                                    best_accuracy = ""
                                try:
                                    best_formatted_address = g.geojson['properties']['address']
                                except:
                                    best_formatted_address = ""                                
                            if confidence_level < 8:
                                is_fail = True
                            else:
                                is_fail = False
                        else:
                            is_fail = True
                        print "confidence level: ", confidence_level
                    else:
                        confidence_level = None
                        bbox = None                            
                        is_fail = True
        
    if is_fail:
        if highest_confidence_level == -1:
            best_g_latlng = "Failed!"

    print "------------- done --------------"
    print "best gocoder: ", best_geocoder
    print "min_bbox: ", min_bbox
         
    t2 = time.time()
    print "================== FINISHED =================="
    print "Time cost: %.4f s" % (t2-t1)
    
    result = {
        'point': best_g_latlng,
        'geocoder': best_geocoder,
        'confidence': highest_confidence_level,
        'accuracy': best_accuracy,
        'formattedaddress': best_formatted_address,
    }
    
    return(result)

@login_required
def instant_geocoding(request):
    address = request.POST["address"]
    print "inst geo for address: ", address
    try:
        print "search in format address"
        get_address = FormattedAddress.objects.get(address=address)
        get_formatted_address = get_address
        point = json.dumps([get_formatted_address.point.lat,get_formatted_address.point.lng])
        geocoder = get_formatted_address.geocoder.name
        confidence = get_formatted_address.confidence_level.score
        response_data = {
            'point': point,
            'geocoder': geocoder,
            'confidence': confidence
        }
    except:
        print "format fail"
        try:
            print "search in address"
            get_address = AddressInventory.objects.get(address=address)
            get_formatted_address = get_address.formatted_address
            point = json.dumps([get_formatted_address.point.lat,get_formatted_address.point.lng])
            geocoder = get_formatted_address.geocoder.name
            confidence = get_formatted_address.confidence_level.score
            response_data = {
                'point': point,
                'geocoder': geocoder,
                'confidence': confidence
            }        
        except:
            print "address fail"
            result = api_geocoding(address,None)
            print "api geoocding result: ", result
            point = json.dumps(result["point"])
            geocoder = result["geocoder"]
            confidence = result["confidence"]
            response_data = {
                'point': point,
                'geocoder': geocoder,
                'confidence': confidence
            }
            print "api response data: ", response_data
            # Google Geocoder Usage
            update_google_geocoder_usage(None,1,True)
            point = result["point"]
            print "point: ",point
            # Save address to address inventory
            if point != "Failed!":
                point = Point(lat=point[0],lng=point[1])
                point.save()
                print "save point"
                geocoder = Geocoder.objects.get(name=result['geocoder'])
                confidence = ConfidenceLevel.objects.get(score=result['confidence'])
                accuracy = result['accuracy']
                formattedaddress = result['formattedaddress']
                print "formatted address: ", formattedaddress
                try:
                    print "try get fortmat"
                    formatted_address = FormattedAddress.objects.get(address=formattedaddress)
                except:
                    print "no format match, save new format address"
                    formatted_address = FormattedAddress(
                        address = formattedaddress,
                        point = point,
                        geocoder = geocoder,
                        confidence_level = confidence,
                        accuracy = accuracy
                    )
                    formatted_address.save()
                print "save new address"
                new_address = AddressInventory(
                    address = address,
                    formatted_address = formatted_address
                )
                new_address.save()

    return HttpResponse(json.dumps(response_data),content_type="application/json")

@login_required
@render_to("geocodingapp/geocoding_setup.html")
def geocoding_setup(request):
    if request.GET["task"]:
        task_id = int(request.GET["task"])
    # read task file
    task = Task.objects.get(id=task_id)
    file_path = task.file.path
    file_ext = file_path[file_path.rfind("."):]
    ## read excel file
    if file_ext in EXCEL_FILE_EXTENSIONS:    
        wb = load_workbook(file_path,read_only=True)
        sheet_names = wb.get_sheet_names()
        ws = wb[sheet_names[0]]
        rownum = ws.get_highest_row()
        colnum = ws.get_highest_column()
        colletter = str(_get_column_letter(colnum))
        preview_range = "A1:%s11" % colletter
        preview_table = []
        for row in ws.iter_rows(preview_range):
            tmp_row = []
            for cell in row:
                if cell.value:
                    tmp_row.append(cell.value)
                else:
                    tmp_row.append("")
            preview_table.append(tmp_row)
    ## read csv file
    elif file_ext in CSV_FILE_EXTENSIONS:
        preview_table = []
        with open(file_path,'rb') as csvfile:
            csvreader = csv.reader(csvfile,delimiter=',',quotechar='"')
            for i in range(11):
                preview_table.append(csvreader.next())
    else:
        print "file not supported"
        preview_table = None
    if preview_table:
        print preview_table
        preview_table_headers = preview_table[0]
        preview_table_content = preview_table[1:]
    else:
        preview_table_headers = None
        preview_table_content = None
    
    return {
            'task_id': task_id,
            'preview_table_headers': preview_table_headers,
            'preview_table_content': preview_table_content
            }

@login_required
def start_geocoding(request):
    is_file_supported = False
    reload(sys)  
    sys.setdefaultencoding('utf8')
    task_id = None
    task = None
    if request.GET["task"]:
        task_id = int(request.GET["task"])
        task = Task.objects.get(id=task_id)
    result_headers = []
    if "label" in request.GET:
        if request.GET["label"] != 'None':
            result_headers.append(request.GET["label"])    
    if "address" in request.GET:
        if request.GET["address"] != 'None':
            result_headers.append(request.GET["address"])
    if "city" in request.GET:
        if request.GET["city"] != 'None':
            result_headers.append(request.GET["city"])
    if "state" in request.GET:
        if request.GET["state"] != 'None':
            result_headers.append(request.GET["state"])            
    if "zip" in request.GET:
        if request.GET["zip"] != 'None':
            result_headers.append(request.GET["zip"])
            
    # read task file
    task = Task.objects.get(id=task_id)
    file_path = task.file.path
    file_ext = file_path[file_path.rfind("."):]
    ## read excel file
    if file_ext in EXCEL_FILE_EXTENSIONS:
        is_file_supported = True
        wb = load_workbook(file_path,read_only=True)
        sheet_names = wb.get_sheet_names()
        ws = wb[sheet_names[0]]
        rownum = ws.get_highest_row()
        colnum = ws.get_highest_column()
        xls_headers = []
        for row in ws.get_squared_range(1,1,colnum,1):
            for cell in row:
                xls_headers.append(cell.value)
        print xls_headers
        result_headers_index = []
        for h in result_headers:
            result_headers_index.append(xls_headers.index(h))    
        print result_headers_index
        result_list = []
        for row in ws.iter_rows(row_offset=1):
            if (type(row[result_headers_index[0]].value) is int) or (type(row[result_headers_index[0]].value) is long):
                label = str(row[result_headers_index[0]].value)
            else:
                if row[result_headers_index[0]].value:
                    label = row[result_headers_index[0]].value.encode('utf-8').strip()
                else:
                    label = ""
            address = u""
            for col in result_headers_index[1:]:
                if row[col].value:
                    if (type(row[col].value) is int) or (type(row[col].value) is long):
                        address += "%s, " % str(row[col].value)
                    else:
                        address += "%s, " % row[col].value.encode('utf-8').strip()
            result_list.append({'label':label,
                                'address':address[:-2]})
    ## read csv file
    elif file_ext in CSV_FILE_EXTENSIONS:
        is_file_supported = True
        result_list = []
        with open(file_path,'rb') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                if (type(row[result_headers[0]]) is int) or (type(row[result_headers[0]]) is long):
                    label = str(row[result_headers[0]])
                else:
                    if row[result_headers[0]]:
                        label = row[result_headers[0]].encode('utf-8').strip()
                    else:
                        label = ""
                address = u""
                for col in result_headers[1:]:
                    if row[col]:
                        if (type(row[col] is int)) or (type(row[col]) is long):
                            address += "%s, " % str(row[col])
                        else:
                            address += "%s, " % row[col].encode('utf-8').strip()
                result_list.append({'label':label,
                                    'address':address[:-2]})
    else:
        print "file type not supported"
        
    if is_file_supported:
        # start geocoding
        for result in result_list:
            address = result['address']        
            # Check if address exists
            print "geo for address: ", address
            try:
                print "search in format address"
                get_address = FormattedAddress.objects.get(address=address)
                get_formatted_address = get_address
                point = [get_formatted_address.point.lat,get_formatted_address.point.lng]
                geocoder = get_formatted_address.geocoder.name
                confidence = get_formatted_address.confidence_level.score
                accuracy = get_formatted_address.accuracy
                geocoding_result = GeocodingResult(
                    task = task,
                    name = result['label'],
                    address = address,
                    formatted_address = get_formatted_address,
                    location = get_formatted_address.point,
                    geocoder = get_formatted_address.geocoder,
                    confidence_level = get_formatted_address.confidence_level,
                    accuracy = accuracy,
                    final_source = FinalSource.objects.get(id=2)
                )
                geocoding_result.save()
            except:
                print "format fail"
                try:
                    print "search in address"
                    get_address = AddressInventory.objects.get(address=address)
                    print "address found in inventory"
                    print get_address
                    get_formatted_address = get_address.formatted_address
                    point = [get_formatted_address.point.lat,get_formatted_address.point.lng]
                    geocoder = get_formatted_address.geocoder.name
                    confidence = get_formatted_address.confidence_level.score
                    accuracy = get_formatted_address.accuracy
                    geocoding_result = GeocodingResult(
                        task = task,
                        name = result['label'],
                        address = address,
                        formatted_address = get_formatted_address,
                        location = get_formatted_address.point,
                        geocoder = get_formatted_address.geocoder,
                        confidence_level = get_formatted_address.confidence_level,
                        accuracy = accuracy,
                        final_source = FinalSource.objects.get(id=2)
                    )
                    geocoding_result.save()      
                except Exception as e:
                    print e
                    print "address fail, use geocoding api"
                    # Create New Google Geocoder Usage
                    google_usage = update_google_geocoder_usage(None,0,True)
                    tmp_result = api_geocoding(address,google_usage.id)
                    tmp_point = tmp_result["point"]
                    if tmp_point == "Failed!":
                        formatted_address = None
                        point = None
                        geocoder = None
                        confidence = ConfidenceLevel.objects.get(score=-1)
                        accuracy = "No matching result found. Geocoding failed!"
                    else:
                        point = Point(lat=tmp_point[0],lng=tmp_point[1])
                        point.save()
                        geocoder = Geocoder.objects.get(name=tmp_result['geocoder'])
                        confidence = ConfidenceLevel.objects.get(score=tmp_result['confidence'])
                        formattedaddress = tmp_result['formattedaddress']
                        accuracy = tmp_result['accuracy']
                        try:
                            formatted_address = FormattedAddress.objects.get(address=formattedaddress)
                        except:
                            formatted_address = FormattedAddress(
                                address = formattedaddress,
                                point = point,
                                geocoder = geocoder,
                                confidence_level = confidence,
                                accuracy = accuracy
                            )
                            formatted_address.save()
                        print "save new address"
                        new_address = AddressInventory(
                            address = address,
                            formatted_address = formatted_address
                        )
                        new_address.save()                    

                        geocoding_result = GeocodingResult(
                            task = task,
                            name = result['label'],
                            address = address,
                            formatted_address = formatted_address,
                            location = point,
                            geocoder = geocoder,
                            confidence_level = confidence,
                            accuracy = accuracy,
                            final_source = FinalSource.objects.get(id=1)
                        )
                        geocoding_result.save()

        task.has_result = True
        task.save()
    
    redirect_url = "%s/geocoding/results?task=%d" % (ROOT_APP_URL,task_id)
    return HttpResponseRedirect(redirect_url)

@login_required
def get_cf_link(request):
    task_id = None
    lat = None
    lng = None
    redirect_url = None
    if "task" in request.GET:
        if request.GET["task"] != 'None':
            task_id = request.GET["task"]
    if "lat" in request.GET:
        if request.GET["lat"] != 'None':
            lat = float(request.GET["lat"])
    if "lng" in request.GET:
        if request.GET["lng"] != 'None':
            lng = float(request.GET["lng"])
    if lat and lng:
        query_str = "Select %s,%s from %s where ST_Contains(geom,ST_Transform(ST_SetSRID(ST_Point(%s,%s),4326),%s))" % (CF_GEOTABLE['col_id'],CF_GEOTABLE['col_name'],CF_GEOTABLE['db_table'],lng,lat,CF_GEOTABLE['srid'])
        cursor = connection.cursor()
        cursor.execute(query_str)
        query_results = cursor.fetchall()
        if query_results:
            nb = query_results[0]
            nb_id = nb[0]
            nb_name = nb[1]
            nb_name_url = ' '.join((re.sub(r'([^\s\w]|_)+','',nb_name)).strip().split()).replace(' ','-').lower()
            redirect_url = "%s%s" % (CF_GEOTABLE['cf_url_nbsummary'],nb_name_url)
        else:
            redirect_url = "%s/geocoding/no_cf_link/%s/" % (ROOT_APP_URL,task_id)
    return HttpResponseRedirect(redirect_url)

@login_required
@render_to("geocodingapp/no_cf_link.html")
def no_cf_link(request,task_id):
    return {"task_id":task_id}

@login_required
@render_to("geocodingapp/geocoding_results.html")
def geocoding_result(request):
    task_id = int(request.GET["task"])
    task = Task.objects.get(id=task_id)
    task_results = GeocodingResult.objects.filter(task=task)
    points = []
    points_wlabels = []
    for result in task_results:
        if result.location:
            points.append([result.location.lat,result.location.lng])
            points_wlabels.append([result.location.lat,result.location.lng,result.name])
    return {"task_id":task_id,"task_results":task_results,"g_points":points,"data_points":json.dumps(points_wlabels).replace("'",r"\'")}


# Export Results as CSV file
def exportcsv_geocodingresults(request,task_id):
    task = Task.objects.get(id=task_id)
    task_results = GeocodingResult.objects.filter(task=task)
    download_data = []
    headers = ("Label","Address","Lat","Long","Geocoder","Confidence Level","Accuracy","Final Source")
    for result in task_results:
        row_data = (
            result.name,
            result.address,
            result.location.lat,
            result.location.lng,
            result.geocoder.name,
            result.confidence_level.score,
            result.accuracy,
            result.final_source.name
        )
        download_data.append(row_data)
       
    # export as CSV
    output_table = StringIO.StringIO()
    csvwriter = csv.writer(output_table,delimiter=",",quotechar='"')
    csvwriter.writerow(headers)
    for row in download_data:
        csvwriter.writerow([unicode(s).encode("utf-8") for s in row])
    tmp_name = "GeocodingTool_results_%s" % datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    outfile = StringIO.StringIO()
    with zipfile.ZipFile(outfile,'w') as zf:
        zf.writestr("geocoding_results.csv",output_table.getvalue())
    zipped_file = outfile.getvalue()
    
    response = HttpResponse(zipped_file, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=%s.zip' % tmp_name

    return response


'''-----------------------
Test functions
-----------------------'''
@login_required
@render_to("geocodingapp/test_geocoding.html")
def test_geocoding(request):
    reload(sys)  
    sys.setdefaultencoding('utf8')    
    if request.GET["task"]:
        task_id = int(request.GET["task"])
    address_headers = []
    if "address" in request.GET:
        if request.GET["address"] != 'None':
            address_headers.append(request.GET["address"])
    if "city" in request.GET:
            if request.GET["city"] != 'None':
                address_headers.append(request.GET["city"])
    if "state" in request.GET:
            if request.GET["state"] != 'None':
                address_headers.append(request.GET["state"])
    if "zip" in request.GET:
            if request.GET["zip"] != 'None':
                address_headers.append(request.GET["zip"])
            
    # read task file
    task = Task.objects.get(id=task_id)
    file_path = task.file.path
    wb = load_workbook(file_path,read_only=True)
    sheet_names = wb.get_sheet_names()
    ws = wb[sheet_names[0]]
    rownum = ws.get_highest_row()
    colnum = ws.get_highest_column()
    xls_headers = []
    for row in ws.get_squared_range(1,1,colnum,1):
        for cell in row:
            xls_headers.append(cell.value)
    print xls_headers
    address_headers_index = []
    for h in address_headers:
        address_headers_index.append(xls_headers.index(h))
    print address_headers_index
    address_list = []
    for row in ws.iter_rows(row_offset=1):
        address = u""
        for col in address_headers_index:
#            print row[col].value
#            print "test"
            if type(row[col].value) is int:
                address += "%s, " % str(row[col].value)
            else:
                address += "%s, " % row[col].value.encode('utf-8').strip()
        address_list.append(address[:-2])
    
    # geocoding
    result_list = []
    failed_list = []
    
    # Order: Google -> OSM -> ArcGIS -> Bing -> Mapbox
    # if not found or confidence lower than 8 (1km), try next
    t1 = time.time()
    for address in address_list:
#        print address
        confidence_level = -1
        highest_confidence_level = -1
        min_bbox = None
        best_geocoder = None
        best_g_latlng = None
        is_fail = False
        # Google geocoder
        print address
        print "Start geocoding, try Google"        
        g = geocoder.google(address)
        print g.geojson
        if g.geojson['properties']['ok']:
            try:
                confidence_level = g.geojson['properties']['confidence']
            except:
                confidence_level = 0
            bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
            if confidence_level > highest_confidence_level:
                highest_confidence_level = confidence_level
                min_bbox = bbox
                print "bbox: ", bbox
                g_latlng = g.geojson['geometry']['coordinates']
                g_latlng.reverse()
                best_g_latlng = g_latlng
                best_geocoder = "Google"
                if confidence_level < 8:
                    is_fail = True
                else:
                    is_fail = False
            else:
                is_fail = True
        else:
            is_fail = True
        if is_fail:        
            print "confidence level: ", confidence_level
            print "google failed, try osm"
            print "---------------"
            confidence_level = None
            bbox = None            
            # OSM geocoder
            g = geocoder.osm(address)
            print g.geojson
            if g.geojson['properties']['ok']:
                try:
                    confidence_level = g.geojson['properties']['confidence']
                except:
                    confidence_level = 0                
                confidence_level = g.geojson['properties']['confidence']
                bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                if confidence_level >= highest_confidence_level:
                    if confidence_level == highest_confidence_level:
                        print "bbox: ", bbox
                        if (not min_bbox) or (bbox < min_bbox):
                            min_bbox = bbox
                            highest_confidence_level = confidence_level
                            g_latlng = g.geojson['geometry']['coordinates']
                            g_latlng.reverse()
                            best_g_latlng = g_latlng
                            best_geocoder = "OSM"
                        else:
                            is_fail = True
                    else:
                        min_bbox = bbox
                        highest_confidence_level = confidence_level
                        g_latlng = g.geojson['geometry']['coordinates']
                        g_latlng.reverse()
                        best_g_latlng = g_latlng
                        best_geocoder = "OSM"
                    if confidence_level < 8:
                        is_fail = True
                    else:
                        is_fail = False
                else:
                    is_fail = True                
            else:              
                is_fail = True
            if is_fail:               
                print "confidence level: ", confidence_level
                print "osm failed, try arcgis"
                print "---------------"   
                confidence_level = None
                bbox = None                              
                # ArcGIS geocoder
                g = geocoder.arcgis(address)
                print g.geojson                
                if g.geojson['properties']['ok']:
                    try:
                        confidence_level = g.geojson['properties']['confidence']
                    except:
                        confidence_level = 0                    
                    confidence_level = g.geojson['properties']['confidence']
                    bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                    if confidence_level >= highest_confidence_level:
                        if confidence_level == highest_confidence_level:
                            print "bbox: ", bbox
                            if (min_bbox) or (bbox > min_bbox):
                                min_bbox = bbox
                                highest_confidence_level = confidence_level
                                g_latlng = g.geojson['geometry']['coordinates']
                                g_latlng.reverse()
                                best_g_latlng = g_latlng
                                best_geocoder = "ArcGIS"
                            else:
                                is_fail = True
                        else:                        
                            highest_confidence_level = confidence_level
                            g_latlng = g.geojson['geometry']['coordinates']
                            g_latlng.reverse()
                            best_g_latlng = g_latlng
                            best_geocoder = "ArcGIS"
                        if confidence_level < 8:
                            is_fail = True
                        else:
                            is_fail = False
                    else:
                        is_fail = True                    
                else:                  
                    is_fail = True
                if is_fail:                   
                    print "confidence level: ", confidence_level
                    print "arcgis failed, try bing"
                    print "---------------"
                    confidence_level = None
                    bbox = None                     
                    # Bing geocoder
                    g = geocoder.bing(address,key=BING_API_KEY)
                    print g.geojson
                    if g.geojson['properties']['ok']:
                        try:
                            confidence_level = g.geojson['properties']['confidence']
                        except:
                            confidence_level = 0                        
                        confidence_level = g.geojson['properties']['confidence']
                        bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                        if confidence_level >= highest_confidence_level:
                            if confidence_level == highest_confidence_level:
                                print "bbox: ", bbox
                                if (min_bbox) or (bbox > min_bbox):
                                    min_bbox = bbox
                                    highest_confidence_level = confidence_level
                                    g_latlng = g.geojson['geometry']['coordinates']
                                    g_latlng.reverse()
                                    best_g_latlng = g_latlng
                                    best_geocoder = "Bing"
                                else:
                                    is_fail = True
                            else:                            
                                highest_confidence_level = confidence_level
                                g_latlng = g.geojson['geometry']['coordinates']
                                g_latlng.reverse()
                                best_g_latlng = g_latlng
                                best_geocoder = "Bing"
                            if confidence_level < 8:
                                is_fail = True
                            else:
                                is_fail = False
                        else:
                            is_fail = True                        
                    else:                      
                        is_fail = True
                    if is_fail:                       
                        print "confidence level: ", confidence_level
                        print "bing failed, try mapbox"
                        print "---------------"
                        confidence_level = None
                        bbox = None                         
                        # Mapbox geocoder
                        g = geocoder.mapbox(address,key=MAPBOX_API_KEY)
                        print "bbox: ", bbox
                        print g.geojson
                        if g.geojson['properties']['ok']:
                            try:
                                confidence_level = g.geojson['properties']['confidence']
                            except:
                                confidence_level = 0                            
                            confidence_level = g.geojson['properties']['confidence']
                            bbox = round(math.sqrt((g.geojson['bbox']['northeast'][0] - g.geojson['bbox']['southwest'][0])**2 + (g.geojson['bbox']['northeast'][1] - g.geojson['bbox']['southwest'][1])**2),8)
                            if confidence_level >= highest_confidence_level:
                                if confidence_level == highest_confidence_level:
                                    print "bbox: ", bbox
                                    if (min_bbox) or (bbox > min_bbox):
                                        min_bbox = bbox
                                        highest_confidence_level = confidence_level
                                        g_latlng = g.geojson['geometry']['coordinates']
                                        g_latlng.reverse()
                                        best_g_latlng = g_latlng
                                        best_geocoder = "Mapbox"
                                    else:
                                        is_fail = True
                                else:                                
                                    highest_confidence_level = confidence_level
                                    g_latlng = g.geojson['geometry']['coordinates']
                                    g_latlng.reverse()
                                    best_g_latlng = g_latlng
                                    best_geocoder = "Mapbox"
                                if confidence_level < 8:
                                    is_fail = True
                                else:
                                    is_fail = False
                            else:
                                is_fail = True
                            print "confidence level: ", confidence_level
                        else:
                            confidence_level = None
                            bbox = None                            
                            is_fail = True
            
        if is_fail:
            if highest_confidence_level > -1:
                result_list.append(best_g_latlng)
            else:
                failed_list.append(address)
        else:
            result_list.append(best_g_latlng)
        print "------------- done --------------"
        print "best gocoder: ", best_geocoder
        print "min_bbox: ", min_bbox
         
    t2 = time.time()
    print "================== FINISHED =================="
    print "Time cost: %.4f s" % (t2-t1)
    return {"g_points":result_list,"failed_list":failed_list}

'''-----------------------
Import Geography
-----------------------'''
GEOGRAPHY_DATA_PATH = STORAGE_ROOTPATH + "geography/"

# Import County
@login_required
def import_geo_county(request):
    try:
        state = State.objects.get(geoid="08")
        with open(GEOGRAPHY_DATA_PATH+"county_state_geography_2010.csv",'rb') as f:
            reader = csv.reader(f)
            for index,row in enumerate(reader):
                if index > 0:
                    county = County(
                        geoid=row[0].strip(),
                        county=row[1].strip(),
                        state=state
                    )
                    county.save()
        return HttpResponse("Geography - County - Import complete!")
    except Exception as e:
        print e
        return HttpResponse("Geography - County - Import Failed! Error: %s" % e)

# Import City
@login_required
def import_geo_city(request):
    try:
        state = State.objects.get(geoid="08")
        with open(GEOGRAPHY_DATA_PATH+"place_state_geography_2010.csv",'rb') as f:
            reader = csv.reader(f)
            for index,row in enumerate(reader):
                if index > 0:
                    try:
                        city = City(
                            geoid=row[0].strip(),
                            city=row[1].strip(),
                            state=state
                        )
                        city.save()
                    except:
                        city = City(
                            geoid=row[0].strip(),
                            city=unicode(row[1].strip(),'latin-1'),
                            state=state
                        )
                        city.save()                     
        return HttpResponse("Geography - City - Import complete!")
    except Exception as e:
        return HttpResponse("Geography - City - Import Failed! Error: %s" % e)
    
    
## Confidence Level
# 10 | less than 0.25 km distance
#  9 | less than 0.5 km distance
#  8 | less than 1 km distance
#  7 | less than 5 km distance
#  6 | less than 7.5 km distance
#  5 | less than 10 km distance
#  4 | less than 15 km distance
#  3 | less than 20 km distance
#  2 | less than 25 km distance
#  1 | 25 km or greater distance
#  0 | unable to determine a bounding box    
    
# return examples
#Google_return =
#{
#	'geometry': {
#		'type': 'Point', 
#		'coordinates': [-104.8348222, 39.7421406]}, 
#	'type': 'Feature', 
#	'properties': {
#		'status': 'OK', 
#		'city': u'Aurora', 
#		'confidence': 9, 
#		'neighborhood': u'Fitzsimons', 
#		'quality': u'street_address', 
#		'encoding': 'utf-8', 
#		'country': u'US', 
#		'provider': 'google', 
#		'location': u'13123 East 16th Avenue, Aurora, CO, 80045', 
#		'county': u'Adams County', 
#		'state': u'CO', 
#		'street': u'E 16th Ave', 
#		'bbox': {'northeast': [39.7434895802915, -104.8334732197085], 'southwest': [39.7407916197085, -104.8361711802915]}, 
#		'status_code': 200, 
#		'address': u'13123 E 16th Ave, Aurora, CO 80045, USA', 
#		'lat': 39.7421406, 
#		'ok': True, 
#		'lng': -104.8348222, 
#		'postal': u'80045', 
#		'housenumber': u'13123', 
#		'accuracy': u'ROOFTOP'}, 
#	'bbox': {'northeast': [39.7434895802915, -104.8334732197085], 'southwest': [39.7407916197085, -104.8361711802915]}
#}
#Google_over_limit_return =
#{
#'type': 'Feature', 
#'properties': {
#    'status': u'OVER_QUERY_LIMIT', 
#    'ok': False, 
#    'encoding': 'utf-8', 
#    'status_code': 200, 
#    'location': u'13123 East 16th Avenue, Aurora, CO, 80045', 
#    'provider': 'google'}
#}

#OSM_return =
#{
#	'geometry': {
#		'type': 'Point', 
#		'coordinates': [-104.834082168753, 39.74226515]}, 
#	'type': 'Feature', 
#	'properties': {
#		'encoding': 'utf-8', 
#		'status_code': 200, 
#		'place_id': u'121500998', 
#		'county': u'Adams County', 
#		'street': u'Donors Way', 
#		'osm_id': u'292811606', 
#		'lng': -104.834082168753, 
#		'quality': u'yes', 
#		'city': u'Aurora', 
#		'confidence': 10, 
#		'type': u'yes', 
#		'state': u'Colorado', 
#		'location': u'13123 East 16th Avenue, Aurora, CO, 80045', 
#		'provider': 'osm', 
#		'housenumber': u'13123',
#		'accuracy': 0.211, 
#		'status': 'OK', 
#		'importance': 0.211, 
#		'bbox': {'northeast': [39.7423792, -104.8336294], 'southwest': [39.7421038, -104.8345339]}, 
#		'address': u'Colorado Institute for Maternal and Fetal Health, 13123, Donors Way, Aurora, Adams County, Colorado, 80045, United States of America', 
#		'lat': 39.74226515, 
#		'postal': u'80045', 
#		'ok': True, 
#		'country': u'United States of America', 
#		'region': u'Colorado', 
#		'osm_type': u'way', 
#		'place_rank': u'30'}, 
#	'bbox': {'northeast': [39.7423792, -104.8336294], 'southwest': [39.7421038, -104.8345339]}
#}
#OSM_result_not_found_return =
#{
#    'type': 'Feature', 
#    'properties': {
#        'status': 'ERROR - No results found', 
#        'ok': False, 
#        'encoding': 'utf-8', 
#        'status_code': 200, 
#        'location': u'12605 E 16th Ave-University of Colorado Hospital, Aurora, CO, 80010', 
#        'provider': 'osm'}
#}

#ArcGIS_return =
#{
#    'geometry': {
#        'type': 'Point', 
#            'coordinates': [-104.86583374892763, 39.74196199605035]}, 
#    'type': 'Feature', 
#    'properties': {
#        'status': 'OK', 
#        'confidence': 6, 
#        'ok': True, 
#        'encoding': 'utf-8', 
#        'status_code': 200, 
#        'provider': 'arcgis', 
#        'score': 94.34, 
#        'bbox': {'northeast': [39.761962, -104.845835], 'southwest': [39.721962, -104.885835]}, 
#        'address': u'E 16th Ave, Aurora, Colorado, 80010', 
#        'lat': 39.74196199605035, 
#        'lng': -104.86583374892763, 
#        'quality': u'StreetName', 
#        'location': u'12605 E 16th Ave-University of Colorado Hospital, Aurora, CO, 80010'}, 
#    'bbox': {'northeast': [39.761962, -104.845835], 'southwest': [39.721962, -104.885835]}
#}

#Bing_no_results_found_return =
#{
#    'type': 'Feature', 
#    'properties': {
#        'status': 'ERROR - No results found', 
#        'ok': False, 
#        'encoding': 'utf-8', 
#        'status_code': 200, 
#        'location': u'12605 E 16th Ave-University of Colorado Hospital, Aurora, CO, 80010', 
#        'provider': 'bing'}
#}

#Mapbox_return = 
#{
#    'geometry': {
#        'type': 'Point', 
#        'coordinates': [-104.841375, 39.742299]}, 
#    'type': 'Feature', 
#    'properties': {
#        'status': 'OK', 
#        'city': u'Aurora', 
#        'confidence': 4, 
#        'ok': True, 
#        'quality': 0.5414545454545455, 
#        'encoding': 'utf-8', 
#        'country': u'United States', 
#        'provider': 'mapbox', 
#        'state': u'Colorado', 
#        'bbox': {'northeast': [39.742880099999994, -104.76411799999997], 'southwest': [39.740920999999986, -104.88312699999997]}, 
#        'status_code': 200, 
#        'address': u'12605 E 16th Ave, Aurora, Colorado 80045, United States', 
#        'lat': 39.742299, 
#        'lng': -104.841375, 
#        'postal': u'80045', 
#        'housenumber': u'12605', 
#        'location': u'12605 E 16th Ave-University of Colorado Hospital, Aurora, CO, 80010'}, 
#    'bbox': {'northeast': [39.742880099999994,-104.76411799999997], 'southwest': [39.740920999999986, -104.88312699999997]}
#}
